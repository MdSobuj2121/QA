import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys  # Import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import datetime

# Set up ChromeDriver automatically
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# Set the path to your Excel file
excel_path = r"C:\Users\user\Desktop\4beatsQ1\4BeatsQ1.xlsx"  # Adjust this path if needed

# Check if the file exists
if not os.path.exists(excel_path):
    print(f"Excel file not found at: {excel_path}. Please check the path.")
    driver.quit()
    exit()

# Load the Excel file
workbook = openpyxl.load_workbook(excel_path)

# Get the current day of the week
today = datetime.datetime.now().strftime('%A')

# Select the sheet based on today's day
if today in workbook.sheetnames:
    sheet = workbook[today]
else:
    print(f"No sheet found for {today}. Please check the Excel file.")
    driver.quit()
    exit()

# Iterate through the rows and get keywords from Column C
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=False):
    keyword = row[0].value
    if not keyword:
        continue  # Skip empty rows

    try:
        # Search the keyword on Google
        driver.get("https://www.google.com")
        
        # Find the search box and enter the keyword
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)

        # Wait for suggestions to load
        wait = WebDriverWait(driver, 10)
        suggestions_element = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'ul.erkvQe'))
        )
        
        # Extract suggestions
        suggestions = suggestions_element.find_elements(By.CSS_SELECTOR, 'li span')
        suggestions_text = [s.text for s in suggestions if s.text]
        
        # Debugging: Print out all suggestions
        print(f"Suggestions for '{keyword}': {suggestions_text}")
        
        # Find the longest and shortest options
        if suggestions_text:
            longest_option = max(suggestions_text, key=len)
            shortest_option = min(suggestions_text, key=len)
            print(f"Keyword: {keyword}, Longest: {longest_option}, Shortest: {shortest_option}")

            # Write results back to Excel
            sheet.cell(row=row[0].row, column=4).value = longest_option  # Column D
            sheet.cell(row=row[0].row, column=5).value = shortest_option  # Column E
        else:
            print(f"No suggestions found for {keyword}.")

    except Exception as e:
        print(f"Error processing keyword '{keyword}': {e}")

# Save the updated Excel file
try:
    workbook.save(excel_path)
    print("Excel file updated successfully.")
except PermissionError:
    print(f"Permission denied: Unable to save the Excel file at {excel_path}. Make sure the file is not open in another program.")

# Close the browser
driver.quit()
